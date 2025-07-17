import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional

def parse_excel_with_categories(file_path: str) -> List[Dict[str, Any]]:
    """
    엑셀 파일의 "Input Sheet"를 파싱하여 발주 데이터를 JSON 리스트로 반환
    대분류, 중분류, 소분류 포함
    
    Args:
        file_path: Excel 파일 경로
        
    Returns:
        List[Dict]: purchase_orders 테이블 구조에 맞는 JSON 리스트
    """
    try:
        # openpyxl을 사용하여 Excel 파일 로드
        workbook = load_workbook(file_path, data_only=True)
        
        # "Input Sheet" 시트 선택
        if "Input Sheet" not in workbook.sheetnames:
            raise ValueError("'Input Sheet' 시트를 찾을 수 없습니다.")
        
        worksheet = workbook["Input Sheet"]
        
        # 발주 데이터 리스트 초기화
        purchase_orders = []
        
        # 2행부터 시작하여 모든 행 읽기
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # 빈 행 건너뛰기 (A열(발주번호)이 비어있는 경우)
            if not row[0]:
                continue
            
            # 각 행을 purchase_orders 구조에 매핑
            # A열부터 순서대로 매핑 (실제 Excel 컬럼 순서에 따라 조정 필요)
            order_data = {
                "order_number": str(row[0]) if row[0] is not None else "",      # A열: 발주번호
                "order_date": format_date(row[1]),                              # B열: 발주일
                "category_lv1": str(row[2]) if row[2] is not None else "",      # C열: 대분류
                "category_lv2": str(row[3]) if row[3] is not None else "",      # D열: 중분류
                "category_lv3": str(row[4]) if row[4] is not None else "",      # E열: 소분류
                "item_name": str(row[5]) if row[5] is not None else "",         # F열: 품목명
                "specification": str(row[6]) if row[6] is not None else "",     # G열: 규격
                "quantity": safe_int(row[7]),                                   # H열: 수량
                "unit_price": safe_int(row[8]),                                 # I열: 단가
                "supply_amount": safe_int(row[9]),                              # J열: 공급가액
                "tax": safe_int(row[10]),                                       # K열: 세액
                "total_amount": safe_int(row[11]),                              # L열: 총금액
                "due_date": format_date(row[12]),                               # M열: 납기일
                "vendor_name": str(row[13]) if len(row) > 13 and row[13] is not None else "",    # N열: 거래처명
                "delivery_name": str(row[14]) if len(row) > 14 and row[14] is not None else "",  # O열: 납품처명
                "note": str(row[15]) if len(row) > 15 and row[15] is not None else ""            # P열: 비고
            }
            
            purchase_orders.append(order_data)
        
        return purchase_orders
        
    except Exception as e:
        print(f"Excel 파싱 중 오류 발생: {str(e)}")
        return []

def parse_excel_with_pandas(file_path: str) -> List[Dict[str, Any]]:
    """
    pandas를 사용한 Excel 파싱 방법
    대분류, 중분류, 소분류 포함
    
    Args:
        file_path: Excel 파일 경로
        
    Returns:
        List[Dict]: purchase_orders 테이블 구조에 맞는 JSON 리스트
    """
    try:
        # pandas로 Excel 파일 읽기 (헤더는 1행)
        df = pd.read_excel(file_path, sheet_name="Input Sheet", header=0)
        
        # 실제 컬럼명을 기준으로 매핑 (Excel 파일의 실제 헤더명에 따라 조정 필요)
        column_mapping = {
            '발주번호': 'order_number',
            '발주일': 'order_date',
            '대분류': 'category_lv1',
            '중분류': 'category_lv2',
            '소분류': 'category_lv3',
            '품목명': 'item_name',
            '규격': 'specification',
            '수량': 'quantity',
            '단가': 'unit_price',
            '공급가액': 'supply_amount',
            '세액': 'tax',
            '총금액': 'total_amount',
            '납기일': 'due_date',
            '거래처명': 'vendor_name',
            '납품처명': 'delivery_name',
            '비고': 'note'
        }
        
        # 컬럼명 변경
        df = df.rename(columns=column_mapping)
        
        # 빈 행 제거
        df = df.dropna(how='all')
        df = df[df['order_number'].notna()]  # 발주번호가 있는 행만 선택
        
        # 데이터 타입 변환 및 JSON 리스트 생성
        purchase_orders = []
        for _, row in df.iterrows():
            order_data = {
                "order_number": str(row.get("order_number", "")) if pd.notna(row.get("order_number")) else "",
                "order_date": format_date(row.get("order_date")),
                "category_lv1": str(row.get("category_lv1", "")) if pd.notna(row.get("category_lv1")) else "",
                "category_lv2": str(row.get("category_lv2", "")) if pd.notna(row.get("category_lv2")) else "",
                "category_lv3": str(row.get("category_lv3", "")) if pd.notna(row.get("category_lv3")) else "",
                "item_name": str(row.get("item_name", "")) if pd.notna(row.get("item_name")) else "",
                "specification": str(row.get("specification", "")) if pd.notna(row.get("specification")) else "",
                "quantity": safe_int(row.get("quantity")),
                "unit_price": safe_int(row.get("unit_price")),
                "supply_amount": safe_int(row.get("supply_amount")),
                "tax": safe_int(row.get("tax")),
                "total_amount": safe_int(row.get("total_amount")),
                "due_date": format_date(row.get("due_date")),
                "vendor_name": str(row.get("vendor_name", "")) if pd.notna(row.get("vendor_name")) else "",
                "delivery_name": str(row.get("delivery_name", "")) if pd.notna(row.get("delivery_name")) else "",
                "note": str(row.get("note", "")) if pd.notna(row.get("note")) else ""
            }
            purchase_orders.append(order_data)
        
        return purchase_orders
        
    except Exception as e:
        print(f"pandas Excel 파싱 중 오류 발생: {str(e)}")
        return []

def format_date(date_value: Any) -> str:
    """
    날짜 값을 YYYY-MM-DD 형식의 문자열로 변환
    
    Args:
        date_value: Excel에서 읽은 날짜 값 (datetime, float, str 등)
        
    Returns:
        str: YYYY-MM-DD 형식의 날짜 문자열
    """
    if date_value is None or pd.isna(date_value):
        return ""
    
    try:
        # datetime 객체인 경우
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        
        # pandas Timestamp인 경우
        if isinstance(date_value, pd.Timestamp):
            return date_value.strftime("%Y-%m-%d")
        
        # Excel 시리얼 번호인 경우 (float)
        if isinstance(date_value, (int, float)):
            # Excel의 1900년 1월 1일을 기준으로 계산
            excel_epoch = datetime(1900, 1, 1)
            # Excel은 1900-01-01을 1로 계산하므로 2를 빼야 함
            # 또한 1900년 2월 29일 버그 처리
            if date_value > 59:
                date_value -= 1
            converted_date = excel_epoch + pd.Timedelta(days=date_value - 2)
            return converted_date.strftime("%Y-%m-%d")
        
        # 문자열인 경우
        if isinstance(date_value, str):
            # 이미 YYYY-MM-DD 형식인지 확인
            if len(date_value) == 10 and date_value.count('-') == 2:
                return date_value
            
            # 다른 형식의 날짜 문자열 파싱 시도
            try:
                parsed_date = pd.to_datetime(date_value)
                return parsed_date.strftime("%Y-%m-%d")
            except:
                return str(date_value)
        
        return str(date_value)
        
    except Exception as e:
        print(f"날짜 변환 중 오류 발생: {date_value} -> {str(e)}")
        return str(date_value) if date_value is not None else ""

def safe_int(value: Any) -> int:
    """
    값을 안전하게 정수로 변환
    
    Args:
        value: 변환할 값
        
    Returns:
        int: 변환된 정수값 (변환 실패시 0)
    """
    if value is None or pd.isna(value):
        return 0
    
    try:
        # 이미 정수인 경우
        if isinstance(value, int):
            return value
        
        # 실수인 경우
        if isinstance(value, float):
            return int(value)
        
        # 문자열인 경우
        if isinstance(value, str):
            # 빈 문자열이나 공백인 경우
            if not value.strip():
                return 0
            
            # 쉼표 제거 후 변환
            cleaned_value = value.replace(",", "").strip()
            return int(float(cleaned_value))
        
        return 0
        
    except Exception as e:
        print(f"숫자 변환 중 오류 발생: {value} -> {str(e)}")
        return 0

# 사용 예시
if __name__ == "__main__":
    # 엑셀 파일 경로
    file_path = "sample_with_categories.xlsx"
    
    print("=== openpyxl 방식으로 파싱 ===")
    result_openpyxl = parse_excel_with_categories(file_path)
    print(json.dumps(result_openpyxl, indent=2, ensure_ascii=False))
    
    print("\n=== pandas 방식으로 파싱 ===")
    result_pandas = parse_excel_with_pandas(file_path)
    print(json.dumps(result_pandas, indent=2, ensure_ascii=False))