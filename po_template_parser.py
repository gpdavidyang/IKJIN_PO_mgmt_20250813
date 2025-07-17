import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional
from collections import defaultdict

def parse_po_template_input(file_path: str) -> Dict[str, Any]:
    """
    PO Template Input 시트를 파싱하여 DB 저장 가능한 형태로 변환
    
    Args:
        file_path: Excel 파일 경로
        
    Returns:
        Dict: 파싱된 데이터 (purchase_orders와 purchase_order_items 분리)
    """
    try:
        # Excel 파일 로드
        workbook = load_workbook(file_path, data_only=True)
        
        if "Input" not in workbook.sheetnames:
            raise ValueError("'Input' 시트를 찾을 수 없습니다.")
        
        worksheet = workbook["Input"]
        
        # 발주서별로 그룹화할 딕셔너리
        orders_by_number = defaultdict(list)
        
        # 2행부터 시작하여 모든 행 읽기
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # 빈 행이거나 발주번호가 없는 경우 건너뛰기
            if not row or not row[0]:
                continue
            
            # 17개 컬럼 (A~Q)만 처리
            row_data = list(row[:17])
            
            # 각 필드 추출
            order_number = str(row_data[0]) if row_data[0] else None
            order_date = format_date(row_data[1])
            site_name = str(row_data[2]) if row_data[2] else ""
            category_lv1 = str(row_data[3]) if row_data[3] else ""
            category_lv2 = str(row_data[4]) if row_data[4] else ""
            category_lv3 = str(row_data[5]) if row_data[5] else ""
            item_name = str(row_data[6]) if row_data[6] else ""
            specification = str(row_data[7]) if row_data[7] else ""
            quantity = safe_number(row_data[8])
            unit_price = safe_number(row_data[9])
            supply_amount = safe_number(row_data[10])
            tax_amount = safe_number(row_data[11])
            total_amount = safe_number(row_data[12])
            due_date = format_date(row_data[13])
            vendor_name = str(row_data[14]) if row_data[14] else ""
            delivery_name = str(row_data[15]) if row_data[15] else ""
            notes = str(row_data[16]) if row_data[16] else ""
            
            # 발주번호가 없으면 건너뛰기
            if not order_number:
                continue
            
            # 발주서 아이템 데이터 생성
            item_data = {
                "itemName": item_name,
                "specification": specification,
                "quantity": quantity,
                "unitPrice": unit_price,
                "supplyAmount": supply_amount,
                "taxAmount": tax_amount,
                "totalAmount": total_amount,
                "categoryLv1": category_lv1,
                "categoryLv2": category_lv2,
                "categoryLv3": category_lv3,
                "deliveryName": delivery_name,
                "notes": notes
            }
            
            # 발주서 정보 (첫 번째 아이템에서 추출)
            order_info = {
                "orderNumber": order_number,
                "orderDate": order_date,
                "siteName": site_name,
                "dueDate": due_date,
                "vendorName": vendor_name
            }
            
            orders_by_number[order_number].append({
                "orderInfo": order_info,
                "itemData": item_data
            })
        
        # 발주서별로 데이터 정리
        parsed_orders = []
        for order_number, items in orders_by_number.items():
            if not items:
                continue
            
            # 첫 번째 아이템에서 발주서 정보 추출
            first_item = items[0]
            order_info = first_item["orderInfo"]
            
            # 모든 아이템 데이터 수집
            item_list = [item["itemData"] for item in items]
            
            # 발주서 총액 계산
            order_total = sum(item["totalAmount"] for item in item_list)
            
            parsed_order = {
                "orderNumber": order_info["orderNumber"],
                "orderDate": order_info["orderDate"],
                "siteName": order_info["siteName"],
                "dueDate": order_info["dueDate"],
                "vendorName": order_info["vendorName"],
                "totalAmount": order_total,
                "items": item_list
            }
            
            parsed_orders.append(parsed_order)
        
        return {
            "success": True,
            "totalOrders": len(parsed_orders),
            "totalItems": sum(len(order["items"]) for order in parsed_orders),
            "orders": parsed_orders
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "orders": []
        }

def format_date(date_value: Any) -> str:
    """날짜 값을 YYYY-MM-DD 형식으로 변환"""
    if date_value is None:
        return ""
    
    try:
        # datetime 객체인 경우
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        
        # Excel 시리얼 번호인 경우
        if isinstance(date_value, (int, float)):
            excel_epoch = datetime(1900, 1, 1)
            if date_value > 59:
                date_value -= 1  # Excel 1900년 버그 보정
            converted_date = excel_epoch + pd.Timedelta(days=date_value - 2)
            return converted_date.strftime("%Y-%m-%d")
        
        # 문자열인 경우
        if isinstance(date_value, str):
            if len(date_value) == 10 and date_value.count('-') == 2:
                return date_value
            try:
                parsed_date = pd.to_datetime(date_value)
                return parsed_date.strftime("%Y-%m-%d")
            except:
                return str(date_value)
        
        return str(date_value)
        
    except Exception as e:
        print(f"날짜 변환 오류: {date_value} -> {str(e)}")
        return str(date_value) if date_value is not None else ""

def safe_number(value: Any) -> float:
    """값을 안전하게 숫자로 변환"""
    if value is None:
        return 0.0
    
    try:
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            if not value.strip():
                return 0.0
            # 쉼표 제거 후 변환
            cleaned_value = value.replace(",", "").strip()
            return float(cleaned_value)
        
        return 0.0
        
    except Exception as e:
        print(f"숫자 변환 오류: {value} -> {str(e)}")
        return 0.0

def extract_sheets_for_email(file_path: str, sheet_names: List[str] = ["갑지", "을지"]) -> Dict[str, Any]:
    """
    이메일 발송을 위한 특정 시트들 추출
    
    Args:
        file_path: Excel 파일 경로
        sheet_names: 추출할 시트 이름들
        
    Returns:
        Dict: 추출된 시트 정보
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        
        extracted_sheets = {}
        for sheet_name in sheet_names:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # 시트 정보 수집
                sheet_info = {
                    "name": sheet_name,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "merged_cells": len(worksheet.merged_cells.ranges),
                    "exists": True
                }
                extracted_sheets[sheet_name] = sheet_info
            else:
                extracted_sheets[sheet_name] = {
                    "name": sheet_name,
                    "exists": False
                }
        
        return {
            "success": True,
            "sheets": extracted_sheets
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "sheets": {}
        }

# 테스트 실행
if __name__ == "__main__":
    file_path = "PO_Template01__Ext_20250716_2.xlsx"
    
    print("=== PO Template Input 시트 파싱 ===")
    result = parse_po_template_input(file_path)
    
    if result["success"]:
        print(f"총 발주서 수: {result['totalOrders']}")
        print(f"총 아이템 수: {result['totalItems']}")
        
        # 첫 번째 발주서 예시 출력
        if result["orders"]:
            first_order = result["orders"][0]
            print(f"\n첫 번째 발주서 예시:")
            print(f"발주번호: {first_order['orderNumber']}")
            print(f"발주일: {first_order['orderDate']}")
            print(f"현장명: {first_order['siteName']}")
            print(f"거래처명: {first_order['vendorName']}")
            print(f"총액: {first_order['totalAmount']:,.0f}원")
            print(f"아이템 수: {len(first_order['items'])}")
            
            if first_order["items"]:
                print(f"첫 번째 아이템: {first_order['items'][0]['itemName']}")
    else:
        print(f"파싱 실패: {result['error']}")
    
    print("\n=== 이메일용 시트 추출 ===")
    sheet_result = extract_sheets_for_email(file_path)
    if sheet_result["success"]:
        for sheet_name, info in sheet_result["sheets"].items():
            if info["exists"]:
                print(f"{sheet_name}: {info['max_row']}행 x {info['max_column']}열")
            else:
                print(f"{sheet_name}: 시트 없음")
    else:
        print(f"시트 추출 실패: {sheet_result['error']}")