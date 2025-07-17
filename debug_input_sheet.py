import pandas as pd
from openpyxl import load_workbook

def debug_input_sheet():
    """Input 시트의 실제 데이터 확인"""
    
    file_path = "PO_Template01__Ext_20250716_2.xlsx"
    
    try:
        # openpyxl로 확인
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook["Input"]
        
        print("=== Input 시트 데이터 디버깅 ===")
        print(f"최대 행: {worksheet.max_row}")
        print(f"최대 열: {worksheet.max_column}")
        
        # 처음 10행 확인
        print("\n처음 10행 확인:")
        for row_num in range(1, min(11, worksheet.max_row + 1)):
            row_data = []
            for col_num in range(1, 18):  # A~Q열
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data.append(str(cell_value)[:20])  # 20자로 제한
                else:
                    row_data.append("None")
            
            print(f"행 {row_num}: {' | '.join(row_data)}")
        
        # 실제 데이터가 있는 행 찾기
        print("\n실제 데이터가 있는 행 찾기:")
        data_rows = []
        for row_num in range(2, worksheet.max_row + 1):
            row_has_data = False
            for col_num in range(1, 18):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None and str(cell_value).strip():
                    row_has_data = True
                    break
            
            if row_has_data:
                data_rows.append(row_num)
                if len(data_rows) >= 5:  # 처음 5개만
                    break
        
        print(f"데이터가 있는 행: {data_rows}")
        
        # 데이터 행 상세 확인
        for row_num in data_rows[:3]:  # 처음 3개 행만
            print(f"\n행 {row_num} 상세:")
            for col_num in range(1, 18):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                col_letter = chr(64 + col_num)
                print(f"  {col_letter}열: {cell_value}")
        
        # pandas로도 확인
        print("\n=== pandas로 확인 ===")
        df = pd.read_excel(file_path, sheet_name="Input", header=0)
        print(f"DataFrame 크기: {df.shape}")
        print(f"비어있지 않은 행: {len(df.dropna(how='all'))}")
        
        # 각 컬럼의 비어있지 않은 값 개수
        print("\n각 컬럼의 데이터 개수:")
        for col in df.columns:
            non_null_count = df[col].notna().sum()
            print(f"  {col}: {non_null_count}개")
        
        # 첫 번째 비어있지 않은 행 찾기
        non_empty_rows = df.dropna(how='all')
        if len(non_empty_rows) > 0:
            print(f"\n첫 번째 비어있지 않은 행:")
            first_row = non_empty_rows.iloc[0]
            for col, value in first_row.items():
                if pd.notna(value):
                    print(f"  {col}: {value}")
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")

if __name__ == "__main__":
    debug_input_sheet()