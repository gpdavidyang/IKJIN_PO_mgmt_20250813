import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_new_sample_excel():
    """새로운 요구사항에 맞는 샘플 Excel 파일 생성"""
    
    # 새로운 컬럼 구조에 맞는 헤더
    headers = [
        "발주번호",      # A열
        "발주일",        # B열
        "품목명",        # C열
        "규격",          # D열
        "수량",          # E열
        "단가",          # F열
        "공급가액",      # G열
        "세액",          # H열
        "총금액",        # I열
        "납기일",        # J열
        "거래처명",      # K열
        "납품처명",      # L열
        "비고"           # M열
    ]
    
    # 샘플 데이터 생성
    sample_data = [
        ["PO-2025-001", "2025-01-15", "LED 조명", "50W", 100, 50000, 4545455, 454545, 5000000, "2025-01-30", "㈜삼성전자", "현대건설 본사", "긴급 주문"],
        ["PO-2025-002", "2025-01-16", "에어컨", "2톤급", 50, 1200000, 54545455, 5454545, 60000000, "2025-02-05", "LG전자", "GS건설 현장사무소", "설치 포함"],
        ["PO-2025-003", "2025-01-17", "철근", "D16", 1000, 8000, 7272727, 727273, 8000000, "2025-01-25", "포스코", "대우건설 자재창고", "품질 검사 필요"],
        ["PO-2025-004", "2025-01-18", "시멘트", "42.5MPa", 200, 15000, 2727273, 272727, 3000000, "2025-02-10", "한국시멘트", "현대건설 현장", "일반 주문"],
        ["PO-2025-005", "2025-01-19", "타일", "300x300", 500, 25000, 11363636, 1136364, 12500000, "2025-02-15", "동양타일", "삼성건설 현장", "고급 타일"]
    ]
    
    # DataFrame 생성
    df = pd.DataFrame(sample_data, columns=headers)
    
    # Excel 파일 생성
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Input Sheet"
    
    # 헤더 추가
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # 데이터 추가
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # 파일 저장
    workbook.save("new_sample.xlsx")
    print("new_sample.xlsx 파일이 생성되었습니다.")
    
    # 구조 확인
    print("\n=== 새로운 샘플 파일 구조 ===")
    for i, header in enumerate(headers):
        print(f"{chr(65+i)}열: {header}")
    
    print(f"\n총 데이터 행 수: {len(sample_data)}")

if __name__ == "__main__":
    create_new_sample_excel()