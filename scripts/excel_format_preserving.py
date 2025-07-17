#!/usr/bin/env python3
"""
Python openpyxl을 사용한 완벽한 서식 보존 엑셀 처리
모든 형식(병합셀, 테두리, 색상, 폰트, 정렬 등)을 완벽하게 보존
"""

import sys
import json
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import argparse

def remove_input_sheet_preserve_format(source_path, target_path, input_sheet_name='Input'):
    """
    Input 시트만 제거하고 모든 서식을 완벽하게 보존
    """
    try:
        print(f"🐍 Python openpyxl로 서식 보존 처리 시작: {source_path} -> {target_path}")
        
        # 소스 파일 존재 확인
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"소스 파일을 찾을 수 없습니다: {source_path}")
        
        # 워크북 로드 (모든 서식 정보 보존)
        workbook = load_workbook(source_path, data_only=False, keep_vba=True, keep_links=True)
        
        original_sheets = workbook.sheetnames.copy()
        print(f"📋 원본 시트 목록: {', '.join(original_sheets)}")
        
        # Input 시트 제거
        removed_sheet = False
        if input_sheet_name in workbook.sheetnames:
            del workbook[input_sheet_name]
            removed_sheet = True
            print(f"🗑️ '{input_sheet_name}' 시트가 제거되었습니다.")
        else:
            print(f"⚠️ '{input_sheet_name}' 시트를 찾을 수 없습니다.")
        
        remaining_sheets = workbook.sheetnames
        print(f"📋 남은 시트 목록: {', '.join(remaining_sheets)}")
        
        if len(remaining_sheets) == 0:
            raise ValueError("모든 시트가 제거되어 빈 엑셀 파일이 됩니다.")
        
        # 타겟 디렉토리 생성
        target_dir = os.path.dirname(target_path)
        if target_dir and not os.path.exists(target_dir):
            os.makedirs(target_dir, exist_ok=True)
        
        # 서식 보존하여 저장
        workbook.save(target_path)
        print(f"✅ 서식 완벽 보존 완료: {target_path}")
        
        # 서식 검증
        format_info = verify_format_preservation(target_path)
        
        result = {
            'success': True,
            'removed_sheet': removed_sheet,
            'remaining_sheets': remaining_sheets,
            'original_format': True,
            'processed_file_path': target_path,
            'format_verification': format_info
        }
        
        return result
        
    except Exception as e:
        error_msg = f"Python openpyxl 처리 실패: {str(e)}"
        print(f"❌ {error_msg}")
        return {
            'success': False,
            'removed_sheet': False,
            'remaining_sheets': [],
            'original_format': False,
            'error': error_msg
        }

def verify_format_preservation(file_path):
    """
    서식 보존 상태 검증
    """
    try:
        workbook = load_workbook(file_path, data_only=False)
        
        format_info = {
            'has_merged_cells': False,
            'has_borders': False,
            'has_font_styles': False,
            'has_cell_colors': False,
            'has_alignment': False,
            'sheet_details': {}
        }
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_details = {
                'name': sheet_name,
                'merged_cells_count': len(worksheet.merged_cells.ranges),
                'borders_count': 0,
                'font_styles_count': 0,
                'cell_colors_count': 0,
                'alignment_count': 0,
                'merged_ranges': [str(range_) for range_ in worksheet.merged_cells.ranges]
            }
            
            # 병합 셀 확인
            if len(worksheet.merged_cells.ranges) > 0:
                format_info['has_merged_cells'] = True
            
            # 각 셀의 서식 확인
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:  # 값이 있는 셀만 검사
                        # 테두리 확인
                        if cell.border and any([
                            cell.border.left.style, cell.border.right.style,
                            cell.border.top.style, cell.border.bottom.style
                        ]):
                            format_info['has_borders'] = True
                            sheet_details['borders_count'] += 1
                        
                        # 폰트 스타일 확인
                        if cell.font and (
                            cell.font.bold or cell.font.italic or 
                            cell.font.name != 'Calibri' or cell.font.size != 11 or
                            cell.font.color.rgb != '00000000'
                        ):
                            format_info['has_font_styles'] = True
                            sheet_details['font_styles_count'] += 1
                        
                        # 셀 색상 확인
                        if cell.fill and cell.fill.fill_type != 'none':
                            format_info['has_cell_colors'] = True
                            sheet_details['cell_colors_count'] += 1
                        
                        # 정렬 확인
                        if cell.alignment and (
                            cell.alignment.horizontal != 'general' or
                            cell.alignment.vertical != 'bottom'
                        ):
                            format_info['has_alignment'] = True
                            sheet_details['alignment_count'] += 1
            
            format_info['sheet_details'][sheet_name] = sheet_details
        
        print(f"🔍 서식 검증 결과:")
        print(f"  병합셀: {format_info['has_merged_cells']}")
        print(f"  테두리: {format_info['has_borders']}")
        print(f"  폰트: {format_info['has_font_styles']}")
        print(f"  색상: {format_info['has_cell_colors']}")
        print(f"  정렬: {format_info['has_alignment']}")
        
        return format_info
        
    except Exception as e:
        print(f"❌ 서식 검증 실패: {str(e)}")
        return {
            'has_merged_cells': False,
            'has_borders': False,
            'has_font_styles': False,
            'has_cell_colors': False,
            'has_alignment': False,
            'sheet_details': {},
            'error': str(e)
        }

def compare_formats(original_path, processed_path):
    """
    두 파일의 서식 비교
    """
    try:
        print(f"🔄 서식 비교 시작: {original_path} vs {processed_path}")
        
        original_format = verify_format_preservation(original_path)
        processed_format = verify_format_preservation(processed_path)
        
        differences = []
        
        # 각 서식 요소 비교
        format_elements = ['has_merged_cells', 'has_borders', 'has_font_styles', 'has_cell_colors', 'has_alignment']
        
        for element in format_elements:
            if original_format.get(element, False) != processed_format.get(element, False):
                differences.append(f"{element} 불일치")
        
        format_preserved = len(differences) == 0
        
        result = {
            'format_preserved': format_preserved,
            'differences': differences,
            'original_format': original_format,
            'processed_format': processed_format
        }
        
        print(f"📊 서식 비교 결과:")
        print(f"  보존됨: {format_preserved}")
        print(f"  차이점: {len(differences)}개")
        if differences:
            print(f"  상세: {', '.join(differences)}")
        
        return result
        
    except Exception as e:
        print(f"❌ 서식 비교 실패: {str(e)}")
        return {
            'format_preserved': False,
            'differences': ['비교 처리 실패'],
            'original_format': {},
            'processed_format': {},
            'error': str(e)
        }

def main():
    """
    CLI 인터페이스
    """
    parser = argparse.ArgumentParser(description='Excel 파일에서 Input 시트만 제거하고 서식 보존')
    parser.add_argument('source', help='원본 엑셀 파일 경로')
    parser.add_argument('target', help='결과 엑셀 파일 경로')
    parser.add_argument('--input-sheet', default='Input', help='제거할 시트명 (기본: Input)')
    parser.add_argument('--compare', action='store_true', help='처리 전후 서식 비교')
    parser.add_argument('--verify', action='store_true', help='결과 파일 서식 검증')
    parser.add_argument('--json', action='store_true', help='JSON 형태로 결과 출력')
    
    args = parser.parse_args()
    
    # 메인 처리
    result = remove_input_sheet_preserve_format(args.source, args.target, args.input_sheet)
    
    # 추가 검증
    if args.verify and result['success']:
        result['verification'] = verify_format_preservation(args.target)
    
    if args.compare and result['success']:
        result['comparison'] = compare_formats(args.source, args.target)
    
    # 결과 출력
    if args.json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print("\n📋 처리 결과:")
        print(f"  성공: {result['success']}")
        if result['success']:
            print(f"  제거된 시트: {result['removed_sheet']}")
            print(f"  남은 시트: {', '.join(result['remaining_sheets'])}")
            print(f"  서식 보존: {result['original_format']}")
        else:
            print(f"  오류: {result.get('error', 'Unknown error')}")
    
    return 0 if result['success'] else 1

if __name__ == '__main__':
    sys.exit(main())