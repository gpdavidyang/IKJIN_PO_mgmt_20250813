import pandas as pd
from openpyxl import load_workbook

def check_excel_structure(file_path: str):
    """Excel 파일의 구조를 확인"""
    
    print("=== Excel 파일 구조 확인 ===")
    
    # openpyxl로 시트 이름 확인
    workbook = load_workbook(file_path, data_only=True)
    print(f"시트 이름들: {workbook.sheetnames}")
    
    # Input Sheet 확인
    if "Input Sheet" in workbook.sheetnames:
        worksheet = workbook["Input Sheet"]
        print(f"\n=== Input Sheet 구조 ===")
        
        # 첫 번째 행 (헤더) 확인
        header_row = []
        for col in range(1, 15):  # A부터 N까지
            cell_value = worksheet.cell(row=1, column=col).value
            header_row.append(cell_value)
        
        print("첫 번째 행 (헤더):")
        for i, header in enumerate(header_row):
            print(f"  {chr(65+i)}열: {header}")
        
        # 두 번째 행 (첫 번째 데이터) 확인
        print("\n두 번째 행 (첫 번째 데이터):")
        data_row = []
        for col in range(1, 15):  # A부터 N까지
            cell_value = worksheet.cell(row=2, column=col).value
            data_row.append(cell_value)
        
        for i, data in enumerate(data_row):
            print(f"  {chr(65+i)}열: {data} (타입: {type(data)})")
        
        # 전체 행 수 확인
        max_row = worksheet.max_row
        print(f"\n전체 행 수: {max_row}")
    
    # pandas로도 확인
    print("\n=== pandas로 확인 ===")
    df = pd.read_excel(file_path, sheet_name="Input Sheet", header=0)
    print(f"컬럼명: {list(df.columns)}")
    print(f"데이터 행 수: {len(df)}")
    print(f"첫 번째 행 데이터:")
    print(df.iloc[0].to_dict())

if __name__ == "__main__":
    check_excel_structure("sample.xlsx")