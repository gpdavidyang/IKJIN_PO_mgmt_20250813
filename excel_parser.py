import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional

def parse_excel_to_purchase_orders(file_path: str) -> List[Dict[str, Any]]:
    """
    Excel 파일의 "Input Sheet"를 파싱하여 purchase_orders 테이블 구조에 맞는 JSON 리스트로 반환
    
    Args:
        file_path: Excel 파일 경로
        
    Returns:
        List[Dict]: purchase_orders 테이블 구조에 맞는 JSON 리스트
    """
    try:
        # openpyxl을 사용하여 Excel 파일 로드
        workbook = load_workbook(file_path, data_only=True)
        
        # "Input Sheet" 시트 선택
        if "Input Sheet" not in workbook.sheetnames:
            raise ValueError("'Input Sheet' 시트를 찾을 수 없습니다.")
        
        worksheet = workbook["Input Sheet"]
        
        # A열부터 M열까지, 2행부터 시작하는 모든 행 읽기
        purchase_orders = []
        
        # 행 순회 (2행부터 시작)
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=13, values_only=True):
            # 빈 행 건너뛰기 (모든 셀이 비어있는 경우)
            if all(cell is None or cell == "" for cell in row):
                continue
            
            # 각 행을 purchase_orders 구조에 매핑
            order_data = {
                "order_number": str(row[0]) if row[0] is not None else "",
                "order_date": format_date(row[1]),
                "item_name": str(row[2]) if row[2] is not None else "",
                "specification": str(row[3]) if row[3] is not None else "",
                "quantity": safe_int(row[4]),
                "unit_price": safe_int(row[5]),
                "supply_amount": safe_int(row[6]),
                "tax": safe_int(row[7]),
                "total_amount": safe_int(row[8]),
                "due_date": format_date(row[9]),
                "vendor_name": str(row[10]) if row[10] is not None else "",
                "delivery_name": str(row[11]) if row[11] is not None else "",
                "note": str(row[12]) if row[12] is not None else ""
            }
            
            purchase_orders.append(order_data)
        
        return purchase_orders
        
    except Exception as e:
        print(f"Excel 파싱 중 오류 발생: {str(e)}")
        return []

def format_date(date_value: Any) -> str:
    """
    날짜 값을 YYYY-MM-DD 형식의 문자열로 변환
    
    Args:
        date_value: Excel에서 읽은 날짜 값 (datetime, float, str 등)
        
    Returns:
        str: YYYY-MM-DD 형식의 날짜 문자열
    """
    if date_value is None:
        return ""
    
    try:
        # datetime 객체인 경우
        if isinstance(date_value, datetime):
            return date_value.strftime("%Y-%m-%d")
        
        # Excel 시리얼 번호인 경우 (float)
        if isinstance(date_value, (int, float)):
            # Excel 시리얼 번호를 datetime으로 변환
            # Excel의 1900년 1월 1일을 기준으로 계산
            excel_epoch = datetime(1900, 1, 1)
            # Excel은 1900-01-01을 1로 계산하므로 2를 빼야 함
            converted_date = excel_epoch + pd.Timedelta(days=date_value - 2)
            return converted_date.strftime("%Y-%m-%d")
        
        # 문자열인 경우
        if isinstance(date_value, str):
            # 이미 YYYY-MM-DD 형식인지 확인
            if len(date_value) == 10 and date_value.count('-') == 2:
                return date_value
            
            # 다른 형식의 날짜 문자열 파싱 시도
            try:
                parsed_date = pd.to_datetime(date_value)
                return parsed_date.strftime("%Y-%m-%d")
            except:
                return str(date_value)
        
        return str(date_value)
        
    except Exception as e:
        print(f"날짜 변환 중 오류 발생: {date_value} -> {str(e)}")
        return str(date_value) if date_value is not None else ""

def safe_int(value: Any) -> int:
    """
    값을 안전하게 정수로 변환
    
    Args:
        value: 변환할 값
        
    Returns:
        int: 변환된 정수값 (변환 실패시 0)
    """
    if value is None:
        return 0
    
    try:
        # 이미 정수인 경우
        if isinstance(value, int):
            return value
        
        # 실수인 경우
        if isinstance(value, float):
            return int(value)
        
        # 문자열인 경우
        if isinstance(value, str):
            # 빈 문자열이나 공백인 경우
            if not value.strip():
                return 0
            
            # 쉼표 제거 후 변환
            cleaned_value = value.replace(",", "").strip()
            return int(float(cleaned_value))
        
        return 0
        
    except Exception as e:
        print(f"숫자 변환 중 오류 발생: {value} -> {str(e)}")
        return 0

def parse_excel_with_pandas(file_path: str) -> List[Dict[str, Any]]:
    """
    pandas를 사용한 대안적 Excel 파싱 방법
    
    Args:
        file_path: Excel 파일 경로
        
    Returns:
        List[Dict]: purchase_orders 테이블 구조에 맞는 JSON 리스트
    """
    try:
        # pandas로 Excel 파일 읽기
        df = pd.read_excel(file_path, sheet_name="Input Sheet", header=None, skiprows=1)
        
        # 컬럼명 설정 (A열부터 M열까지)
        column_names = [
            "order_number", "order_date", "item_name", "specification",
            "quantity", "unit_price", "supply_amount", "tax", "total_amount",
            "due_date", "vendor_name", "delivery_name", "note"
        ]
        
        # 13개 컬럼만 선택
        df = df.iloc[:, :13]
        df.columns = column_names
        
        # 빈 행 제거
        df = df.dropna(how='all')
        
        # 데이터 타입 변환
        purchase_orders = []
        for _, row in df.iterrows():
            order_data = {
                "order_number": str(row["order_number"]) if pd.notna(row["order_number"]) else "",
                "order_date": format_date(row["order_date"]),
                "item_name": str(row["item_name"]) if pd.notna(row["item_name"]) else "",
                "specification": str(row["specification"]) if pd.notna(row["specification"]) else "",
                "quantity": safe_int(row["quantity"]),
                "unit_price": safe_int(row["unit_price"]),
                "supply_amount": safe_int(row["supply_amount"]),
                "tax": safe_int(row["tax"]),
                "total_amount": safe_int(row["total_amount"]),
                "due_date": format_date(row["due_date"]),
                "vendor_name": str(row["vendor_name"]) if pd.notna(row["vendor_name"]) else "",
                "delivery_name": str(row["delivery_name"]) if pd.notna(row["delivery_name"]) else "",
                "note": str(row["note"]) if pd.notna(row["note"]) else ""
            }
            purchase_orders.append(order_data)
        
        return purchase_orders
        
    except Exception as e:
        print(f"pandas Excel 파싱 중 오류 발생: {str(e)}")
        return []

# 테스트 실행
if __name__ == "__main__":
    # 샘플 Excel 파일로 테스트
    file_path = "sample.xlsx"
    
    print("=== openpyxl 방식으로 파싱 ===")
    result_openpyxl = parse_excel_to_purchase_orders(file_path)
    print(json.dumps(result_openpyxl, indent=2, ensure_ascii=False))
    
    print("\n=== pandas 방식으로 파싱 ===")
    result_pandas = parse_excel_with_pandas(file_path)
    print(json.dumps(result_pandas, indent=2, ensure_ascii=False))