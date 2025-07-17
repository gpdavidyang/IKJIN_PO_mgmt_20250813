#!/usr/bin/env python3
"""
최소한의 처리로 Input 시트만 삭제하는 스크립트
원본 파일을 복사한 후 Input 시트만 삭제하여 서식 완전 보존
"""

import sys
import os
import json
import shutil
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def remove_input_sheet_minimal(source_path, target_path, input_sheet_name='Input'):
    """
    최소한의 처리로 Input 시트만 삭제
    1. 원본 파일을 타겟 경로로 복사
    2. 복사된 파일에서 Input 시트만 삭제
    3. 저장 (다른 시트들은 전혀 건드리지 않음)
    """
    result = {
        'success': False,
        'removed_sheet': False,
        'remaining_sheets': [],
        'original_format': True,
        'error': None,
        'method': 'minimal_processing'
    }
    
    try:
        print(f"📋 최소한의 처리 시작: {source_path} -> {target_path}", file=sys.stderr)
        
        # 소스 파일 존재 확인
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"소스 파일을 찾을 수 없습니다: {source_path}")
        
        # 1단계: 원본 파일을 타겟 경로로 완전 복사
        shutil.copy2(source_path, target_path)
        print(f"✅ 원본 파일 완전 복사 완료", file=sys.stderr)
        
        # 2단계: 복사된 파일에서 Input 시트만 삭제
        workbook = load_workbook(
            target_path,
            read_only=False,
            keep_vba=True,
            keep_links=True,
            data_only=False
        )
        
        # 원본 시트 목록
        original_sheets = workbook.sheetnames.copy()
        print(f"📋 원본 시트 목록: {', '.join(original_sheets)}", file=sys.stderr)
        
        # Input 시트 찾기 및 삭제
        removed_sheet = False
        if input_sheet_name in workbook.sheetnames:
            # Input 시트만 삭제 (다른 시트는 전혀 건드리지 않음)
            workbook.remove(workbook[input_sheet_name])
            removed_sheet = True
            print(f"🗑️ '{input_sheet_name}' 시트만 삭제됨", file=sys.stderr)
        else:
            print(f"⚠️ '{input_sheet_name}' 시트를 찾을 수 없습니다.", file=sys.stderr)
        
        # 남은 시트 목록
        remaining_sheets = workbook.sheetnames.copy()
        print(f"📋 남은 시트 목록: {', '.join(remaining_sheets)}", file=sys.stderr)
        
        if len(remaining_sheets) == 0:
            raise ValueError("모든 시트가 제거되어 빈 엑셀 파일이 됩니다.")
        
        # 3단계: 최소한의 저장 (원본 서식 유지)
        workbook.save(target_path)
        workbook.close()
        
        print(f"✅ 최소한의 처리 완료 (원본 서식 완전 보존)", file=sys.stderr)
        
        result.update({
            'success': True,
            'removed_sheet': removed_sheet,
            'remaining_sheets': remaining_sheets,
            'original_format': True
        })
        
    except InvalidFileException as e:
        error_msg = f"올바른 엑셀 파일이 아닙니다: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
    except FileNotFoundError as e:
        error_msg = f"파일을 찾을 수 없습니다: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
    except Exception as e:
        error_msg = f"최소한의 처리 실패: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
        # 실패 시 타겟 파일 삭제
        if os.path.exists(target_path):
            os.remove(target_path)
            print(f"🗑️ 실패한 타겟 파일 삭제: {target_path}", file=sys.stderr)
    
    return result

def copy_file_and_remove_sheet_binary(source_path, target_path, input_sheet_name='Input'):
    """
    바이너리 레벨에서 파일 복사 후 Input 시트만 제거
    """
    result = {
        'success': False,
        'removed_sheet': False,
        'remaining_sheets': [],
        'original_format': True,
        'error': None,
        'method': 'binary_copy'
    }
    
    try:
        print(f"🔧 바이너리 복사 후 처리 시작: {source_path} -> {target_path}", file=sys.stderr)
        
        # 소스 파일 존재 확인
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"소스 파일을 찾을 수 없습니다: {source_path}")
        
        # 1단계: 바이너리 레벨에서 완전 복사
        with open(source_path, 'rb') as src_file:
            with open(target_path, 'wb') as dst_file:
                shutil.copyfileobj(src_file, dst_file)
        
        print(f"✅ 바이너리 복사 완료", file=sys.stderr)
        
        # 2단계: 복사된 파일에서 Input 시트만 삭제
        workbook = load_workbook(
            target_path,
            read_only=False,
            keep_vba=True,
            keep_links=True,
            data_only=False
        )
        
        # 원본 시트 목록
        original_sheets = workbook.sheetnames.copy()
        print(f"📋 시트 목록: {', '.join(original_sheets)}", file=sys.stderr)
        
        # Input 시트 찾기 및 삭제
        removed_sheet = False
        if input_sheet_name in workbook.sheetnames:
            workbook.remove(workbook[input_sheet_name])
            removed_sheet = True
            print(f"🗑️ '{input_sheet_name}' 시트만 삭제됨", file=sys.stderr)
        else:
            print(f"⚠️ '{input_sheet_name}' 시트를 찾을 수 없습니다.", file=sys.stderr)
        
        # 남은 시트 목록
        remaining_sheets = workbook.sheetnames.copy()
        print(f"📋 남은 시트 목록: {', '.join(remaining_sheets)}", file=sys.stderr)
        
        if len(remaining_sheets) == 0:
            raise ValueError("모든 시트가 제거되어 빈 엑셀 파일이 됩니다.")
        
        # 3단계: 저장
        workbook.save(target_path)
        workbook.close()
        
        print(f"✅ 바이너리 복사 후 처리 완료", file=sys.stderr)
        
        result.update({
            'success': True,
            'removed_sheet': removed_sheet,
            'remaining_sheets': remaining_sheets,
            'original_format': True
        })
        
    except Exception as e:
        error_msg = f"바이너리 복사 후 처리 실패: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
        # 실패 시 타겟 파일 삭제
        if os.path.exists(target_path):
            os.remove(target_path)
            print(f"🗑️ 실패한 타겟 파일 삭제: {target_path}", file=sys.stderr)
    
    return result

def main():
    """
    메인 함수 - 커맨드라인 인자 처리
    """
    if len(sys.argv) < 2:
        print("사용법: python excel-minimal-processing.py <command> [args...]", file=sys.stderr)
        print("  minimal <source> <target> [sheet_name]: 최소한의 처리로 Input 시트 제거", file=sys.stderr)
        print("  binary <source> <target> [sheet_name]: 바이너리 복사 후 Input 시트 제거", file=sys.stderr)
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'minimal':
        if len(sys.argv) != 5:
            print("사용법: python excel-minimal-processing.py minimal <source_path> <target_path> <input_sheet_name>", file=sys.stderr)
            sys.exit(1)
        
        source_path = sys.argv[2]
        target_path = sys.argv[3]
        input_sheet_name = sys.argv[4]
        
        # 최소한의 처리 실행
        result = remove_input_sheet_minimal(source_path, target_path, input_sheet_name)
        
        # 결과를 JSON으로 출력
        print(json.dumps(result, ensure_ascii=False, indent=2))
        
        # 성공/실패에 따른 exit code
        sys.exit(0 if result['success'] else 1)
        
    elif command == 'binary':
        if len(sys.argv) != 5:
            print("사용법: python excel-minimal-processing.py binary <source_path> <target_path> <input_sheet_name>", file=sys.stderr)
            sys.exit(1)
        
        source_path = sys.argv[2]
        target_path = sys.argv[3]
        input_sheet_name = sys.argv[4]
        
        # 바이너리 복사 후 처리 실행
        result = copy_file_and_remove_sheet_binary(source_path, target_path, input_sheet_name)
        
        # 결과를 JSON으로 출력
        print(json.dumps(result, ensure_ascii=False, indent=2))
        
        # 성공/실패에 따른 exit code
        sys.exit(0 if result['success'] else 1)
    
    else:
        print(f"알 수 없는 명령: {command}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()