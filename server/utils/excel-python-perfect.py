#!/usr/bin/env python3
"""
Python openpyxl을 사용한 완벽한 엑셀 서식 보존 처리
이 스크립트는 Node.js에서 호출되어 Input 시트만 제거하고 모든 서식을 보존합니다.
"""

import sys
import os
import json
import shutil
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def remove_input_sheet_perfect(source_path, target_path, input_sheet_name='Input'):
    """
    openpyxl을 사용하여 Input 시트만 제거하고 모든 서식 보존
    """
    result = {
        'success': False,
        'removed_sheet': False,
        'remaining_sheets': [],
        'original_format': True,
        'error': None
    }
    
    try:
        print(f"🐍 Python openpyxl 처리 시작: {source_path} -> {target_path}", file=sys.stderr)
        
        # 소스 파일 존재 확인
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"소스 파일을 찾을 수 없습니다: {source_path}")
        
        # 1단계: 원본 파일을 타겟 경로로 복사
        shutil.copy2(source_path, target_path)
        print(f"✅ 원본 파일 복사 완료", file=sys.stderr)
        
        # 2단계: 복사된 파일을 openpyxl로 로드
        # keep_vba=True, keep_links=True로 모든 정보 보존
        workbook = load_workbook(
            target_path,
            read_only=False,
            keep_vba=True,
            keep_links=True,
            data_only=False
        )
        
        # 원본 시트 목록
        original_sheets = workbook.sheetnames.copy()
        print(f"📋 원본 시트 목록: {', '.join(original_sheets)}", file=sys.stderr)
        
        # Input 시트 찾기 및 제거
        removed_sheet = False
        if input_sheet_name in workbook.sheetnames:
            # Input 시트 제거
            workbook.remove(workbook[input_sheet_name])
            removed_sheet = True
            print(f"🗑️ '{input_sheet_name}' 시트가 제거되었습니다.", file=sys.stderr)
        else:
            print(f"⚠️ '{input_sheet_name}' 시트를 찾을 수 없습니다.", file=sys.stderr)
        
        # 남은 시트 목록
        remaining_sheets = workbook.sheetnames.copy()
        print(f"📋 남은 시트 목록: {', '.join(remaining_sheets)}", file=sys.stderr)
        
        if len(remaining_sheets) == 0:
            raise ValueError("모든 시트가 제거되어 빈 엑셀 파일이 됩니다.")
        
        # 3단계: 수정된 파일 저장
        # 모든 서식과 스타일 정보 보존
        workbook.save(target_path)
        workbook.close()
        
        print(f"✅ Python openpyxl 처리 완료 (완벽한 서식 보존)", file=sys.stderr)
        
        result.update({
            'success': True,
            'removed_sheet': removed_sheet,
            'remaining_sheets': remaining_sheets,
            'original_format': True
        })
        
    except InvalidFileException as e:
        error_msg = f"올바른 엑셀 파일이 아닙니다: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
    except FileNotFoundError as e:
        error_msg = f"파일을 찾을 수 없습니다: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
    except Exception as e:
        error_msg = f"Python 처리 실패: {str(e)}"
        print(f"❌ {error_msg}", file=sys.stderr)
        result.update({
            'success': False,
            'error': error_msg
        })
        
        # 실패 시 타겟 파일 삭제
        if os.path.exists(target_path):
            os.remove(target_path)
            print(f"🗑️ 실패한 타겟 파일 삭제: {target_path}", file=sys.stderr)
    
    return result

def main():
    """
    메인 함수 - 커맨드라인 인자 처리
    """
    if len(sys.argv) != 4:
        print("사용법: python excel-python-perfect.py <source_path> <target_path> <input_sheet_name>", file=sys.stderr)
        sys.exit(1)
    
    source_path = sys.argv[1]
    target_path = sys.argv[2]
    input_sheet_name = sys.argv[3]
    
    # 처리 실행
    result = remove_input_sheet_perfect(source_path, target_path, input_sheet_name)
    
    # 결과를 JSON으로 출력 (Node.js가 읽을 수 있도록)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    # 성공/실패에 따른 exit code
    sys.exit(0 if result['success'] else 1)

if __name__ == "__main__":
    main()