import pandas as pd
from openpyxl import load_workbook
import json

def analyze_po_template():
    """PO_Template01_Ext_20250716_2.xlsx 파일의 구조를 분석"""
    
    file_path = "PO_Template01__Ext_20250716_2.xlsx"
    
    try:
        # 워크북 로드
        workbook = load_workbook(file_path, data_only=True)
        print(f"=== {file_path} 파일 분석 ===")
        print(f"시트 이름들: {workbook.sheetnames}")
        
        # Input 시트 분석
        if "Input" in workbook.sheetnames:
            print("\n=== Input 시트 분석 ===")
            worksheet = workbook["Input"]
            
            # 헤더 확인 (첫 번째 행)
            print("헤더 정보:")
            headers = []
            for col in range(1, 20):  # A부터 S까지 확인
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(cell_value)
                    print(f"  {chr(64+col)}열: {cell_value}")
                else:
                    break
            
            # 데이터 샘플 확인
            print(f"\n데이터 행 수: {worksheet.max_row - 1}")
            if worksheet.max_row > 1:
                print("\n첫 번째 데이터 행:")
                for col in range(1, len(headers) + 1):
                    cell_value = worksheet.cell(row=2, column=col).value
                    print(f"  {chr(64+col)}열 ({headers[col-1]}): {cell_value}")
        
        # 갑지 시트 분석
        if "갑지" in workbook.sheetnames:
            print("\n=== 갑지 시트 분석 ===")
            worksheet = workbook["갑지"]
            print(f"갑지 시트 크기: {worksheet.max_row}행 x {worksheet.max_column}열")
            
            # 병합된 셀이나 특별한 구조 확인
            merged_ranges = worksheet.merged_cells.ranges
            if merged_ranges:
                print(f"병합된 셀 범위: {len(merged_ranges)}개")
                for i, merged_range in enumerate(merged_ranges):
                    if i < 5:  # 처음 5개만 출력
                        print(f"  {merged_range}")
        
        # 을지 시트 분석
        if "을지" in workbook.sheetnames:
            print("\n=== 을지 시트 분석 ===")
            worksheet = workbook["을지"]
            print(f"을지 시트 크기: {worksheet.max_row}행 x {worksheet.max_column}열")
            
            # 병합된 셀이나 특별한 구조 확인
            merged_ranges = worksheet.merged_cells.ranges
            if merged_ranges:
                print(f"병합된 셀 범위: {len(merged_ranges)}개")
                for i, merged_range in enumerate(merged_ranges):
                    if i < 5:  # 처음 5개만 출력
                        print(f"  {merged_range}")
        
        # pandas로도 Input 시트 확인
        print("\n=== pandas로 Input 시트 확인 ===")
        df = pd.read_excel(file_path, sheet_name="Input")
        print(f"컬럼명: {list(df.columns)}")
        print(f"데이터 행 수: {len(df)}")
        
        if len(df) > 0:
            print("첫 번째 행 데이터:")
            first_row = df.iloc[0].to_dict()
            for key, value in first_row.items():
                print(f"  {key}: {value}")
        
        return {
            "sheets": workbook.sheetnames,
            "input_headers": headers if 'headers' in locals() else [],
            "input_data_rows": worksheet.max_row - 1 if 'Input' in workbook.sheetnames else 0
        }
        
    except Exception as e:
        print(f"파일 분석 중 오류 발생: {str(e)}")
        return None

if __name__ == "__main__":
    result = analyze_po_template()
    if result:
        print(f"\n=== 분석 결과 요약 ===")
        print(json.dumps(result, indent=2, ensure_ascii=False))