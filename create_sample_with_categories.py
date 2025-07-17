import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_sample_excel_with_categories():
    """분류 항목이 포함된 샘플 Excel 파일 생성"""
    
    # 헤더 정의 (분류 항목 포함)
    headers = [
        "발주번호",      # A열
        "발주일",        # B열
        "대분류",        # C열
        "중분류",        # D열
        "소분류",        # E열
        "품목명",        # F열
        "규격",          # G열
        "수량",          # H열
        "단가",          # I열
        "공급가액",      # J열
        "세액",          # K열
        "총금액",        # L열
        "납기일",        # M열
        "거래처명",      # N열
        "납품처명",      # O열
        "비고"           # P열
    ]
    
    # 샘플 데이터 생성 (분류 항목 포함)
    sample_data = [
        ["PO-2025-001", "2025-01-15", "전기/전자", "조명", "LED", "LED 조명", "50W", 100, 50000, 4545455, 454545, 5000000, "2025-01-30", "㈜삼성전자", "현대건설 본사", "긴급 주문"],
        ["PO-2025-002", "2025-01-16", "전기/전자", "냉난방", "에어컨", "에어컨", "2톤급", 50, 1200000, 54545455, 5454545, 60000000, "2025-02-05", "LG전자", "GS건설 현장사무소", "설치 포함"],
        ["PO-2025-003", "2025-01-17", "건축자재", "철강", "철근", "철근", "D16", 1000, 8000, 7272727, 727273, 8000000, "2025-01-25", "포스코", "대우건설 자재창고", "품질 검사 필요"],
        ["PO-2025-004", "2025-01-18", "건축자재", "시멘트", "일반시멘트", "시멘트", "42.5MPa", 200, 15000, 2727273, 272727, 3000000, "2025-02-10", "한국시멘트", "현대건설 현장", "일반 주문"],
        ["PO-2025-005", "2025-01-19", "건축자재", "마감재", "타일", "타일", "300x300", 500, 25000, 11363636, 1136364, 12500000, "2025-02-15", "동양타일", "삼성건설 현장", "고급 타일"],
        ["PO-2025-006", "2025-01-20", "설비", "배관", "PVC파이프", "배관 파이프", "100mm", 300, 8000, 2181818, 218182, 2400000, "2025-02-01", "대성파이프", "롯데건설 현장", "KS인증 제품"],
        ["PO-2025-007", "2025-01-21", "전기/전자", "통신", "인터넷 장비", "라우터", "기가비트", 20, 150000, 2727273, 272727, 3000000, "2025-02-12", "SK네트웍스", "KT 통신센터", "보안 인증 필수"],
        ["PO-2025-008", "2025-01-22", "안전용품", "보호구", "안전모", "안전 헬멧", "일반형", 200, 15000, 2727273, 272727, 3000000, "2025-01-28", "안전용품센터", "대림산업 현장", "산업안전 규격"],
        ["PO-2025-009", "2025-01-23", "사무용품", "가구", "책상", "사무용 책상", "1600x800", 30, 200000, 5454545, 545455, 6000000, "2025-02-20", "한샘가구", "포스코건설 사무실", "조립 배송"],
        ["PO-2025-010", "2025-01-24", "건축자재", "목재", "합판", "합판", "18T", 100, 35000, 3181818, 318182, 3500000, "2025-02-08", "동화기업", "현대엔지니어링", "방수 처리"]
    ]
    
    # Workbook 생성
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Input Sheet"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 추가
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 추가
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # 컬럼 너비 자동 조정
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 추가 시트 생성 (갑지, 을지)
    worksheet2 = workbook.create_sheet("갑지")
    worksheet2.cell(row=1, column=1, value="갑지 시트 (추후 구현)")
    
    worksheet3 = workbook.create_sheet("을지")
    worksheet3.cell(row=1, column=1, value="을지 시트 (추후 구현)")
    
    # 파일 저장
    workbook.save("sample_with_categories.xlsx")
    print("sample_with_categories.xlsx 파일이 생성되었습니다.")
    
    # 구조 확인 출력
    print("\n=== 새로운 샘플 파일 구조 ===")
    print("컬럼 구조:")
    for i, header in enumerate(headers):
        print(f"  {chr(65+i)}열: {header}")
    
    print(f"\n총 데이터 행 수: {len(sample_data)}")
    print("\n분류 체계 예시:")
    print("  - 전기/전자 > 조명 > LED")
    print("  - 건축자재 > 철강 > 철근")
    print("  - 설비 > 배관 > PVC파이프")
    print("  - 안전용품 > 보호구 > 안전모")
    print("  - 사무용품 > 가구 > 책상")

if __name__ == "__main__":
    create_sample_excel_with_categories()