#!/usr/bin/env python3
"""
Update column names in all test Excel files to match standard template
"""
import os
from openpyxl import load_workbook
import glob

# Define the column mapping from old to new format
COLUMN_MAPPING = {
    '거래처명': '거래처명',  # Keep as is
    '거래처': '거래처명',     # Map to standard
    '현장명': '프로젝트명',    # Change to standard
    '발주일': '발주일자',      # Add '자' suffix
    '납기일': '납기일자',      # Add '자' suffix
    '발주번호': None,         # Remove this column (not in template)
    '품목': '품목명',         # Add '명' suffix
    '규격': '규격',          # Keep as is
    '수량': '수량',          # Keep as is
    '단위': None,            # Remove this column (not in template)
    '단가': '단가',          # Keep as is
    '공급가액': None,         # Remove (will calculate 총금액)
    '부가세': None,          # Remove (will calculate 총금액)
    '합계': '총금액',         # Change to standard
    '대분류': '대분류',       # Keep as is
    '중분류': '중분류',       # Keep as is
    '소분류': '소분류',       # Keep as is
    '비고': '비고'           # Keep as is
}

# Standard template column order (based on PO_Excel_Template.xlsx)
STANDARD_COLUMNS = [
    '발주일자',
    '납기일자', 
    '거래처명',
    '거래처 이메일',
    '납품처명',
    '납품처 이메일',
    '프로젝트명',
    '대분류',
    '중분류',
    '소분류',
    '품목명',
    '규격',
    '수량',
    '단가',
    '총금액',
    '비고'
]

def update_excel_file(file_path):
    """Update column names in a single Excel file"""
    print(f"Processing: {os.path.basename(file_path)}")
    
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Check if Input sheet exists
        if 'Input' not in wb.sheetnames:
            print(f"  ⚠️  No 'Input' sheet found in {file_path}")
            return False
            
        # Get the Input sheet
        sheet = wb['Input']
        
        # Read current headers (first row)
        current_headers = []
        for cell in sheet[1]:
            current_headers.append(cell.value)
        
        print(f"  Current headers: {current_headers[:5]}...")  # Show first 5
        
        # Create a mapping for data preservation
        data_mapping = {}
        for col_idx, old_header in enumerate(current_headers, 1):
            if old_header and old_header in COLUMN_MAPPING:
                new_header = COLUMN_MAPPING[old_header]
                if new_header:  # Only map if not None (not removed)
                    data_mapping[col_idx] = new_header
        
        # Read all data rows (skip header)
        data_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                data_rows.append(row)
        
        # Clear the sheet
        sheet.delete_rows(1, sheet.max_row)
        
        # Write new headers in standard order
        for col_idx, header in enumerate(STANDARD_COLUMNS, 1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Write data back with proper mapping
        for row_idx, old_row in enumerate(data_rows, 2):
            for old_col_idx, old_value in enumerate(old_row, 1):
                if old_col_idx in data_mapping:
                    new_header = data_mapping[old_col_idx]
                    if new_header in STANDARD_COLUMNS:
                        new_col_idx = STANDARD_COLUMNS.index(new_header) + 1
                        
                        # Special handling for date columns
                        if new_header in ['발주일자', '납기일자'] and old_value:
                            # Keep the date value as is
                            sheet.cell(row=row_idx, column=new_col_idx, value=old_value)
                        # Special handling for 총금액 (if it was 합계)
                        elif new_header == '총금액' and old_value:
                            sheet.cell(row=row_idx, column=new_col_idx, value=old_value)
                        else:
                            sheet.cell(row=row_idx, column=new_col_idx, value=old_value)
                
                # Add empty email fields if needed
                email_fields = ['거래처 이메일', '납품처 이메일', '납품처명']
                for field in email_fields:
                    if field in STANDARD_COLUMNS:
                        col_idx = STANDARD_COLUMNS.index(field) + 1
                        current_value = sheet.cell(row=row_idx, column=col_idx).value
                        if not current_value:
                            if field == '거래처 이메일':
                                # Get vendor name
                                vendor_col = STANDARD_COLUMNS.index('거래처명') + 1
                                vendor_name = sheet.cell(row=row_idx, column=vendor_col).value
                                if vendor_name:
                                    sheet.cell(row=row_idx, column=col_idx, value=f"{vendor_name}@example.com")
                            elif field == '납품처명':
                                # Use project name as delivery location
                                project_col = STANDARD_COLUMNS.index('프로젝트명') + 1
                                project_name = sheet.cell(row=row_idx, column=project_col).value
                                if project_name:
                                    sheet.cell(row=row_idx, column=col_idx, value=project_name)
                            elif field == '납품처 이메일':
                                sheet.cell(row=row_idx, column=col_idx, value="delivery@example.com")
        
        # Save the workbook
        wb.save(file_path)
        print(f"  ✅ Updated successfully")
        return True
        
    except Exception as e:
        print(f"  ❌ Error: {e}")
        return False

def main():
    """Main function to update all test files"""
    test_dir = "/Users/david/workspace/IKJIN_PO_20250826/IKJIN_PO_mgmt_20250813/PO_test/generated_test_files"
    
    # Get all Excel files
    excel_files = glob.glob(os.path.join(test_dir, "*.xlsx"))
    
    print(f"Found {len(excel_files)} Excel files to update\n")
    
    success_count = 0
    failed_files = []
    
    for file_path in excel_files:
        if update_excel_file(file_path):
            success_count += 1
        else:
            failed_files.append(os.path.basename(file_path))
    
    print(f"\n{'='*50}")
    print(f"Summary:")
    print(f"  Total files: {len(excel_files)}")
    print(f"  Successfully updated: {success_count}")
    print(f"  Failed: {len(failed_files)}")
    
    if failed_files:
        print(f"\nFailed files:")
        for file in failed_files:
            print(f"  - {file}")
    
    print(f"\n✅ Column name update completed!")

if __name__ == "__main__":
    main()